#!/usr/bin/env python3
"""Parse ADP Rankings/ into per-bucket master lists (one column per site).

Writes:
  rankings/adp-sources/<bucket>.csv
  rankings/adp-sources/master.xlsx
  rankings/adp-sources/_report.txt
"""

from __future__ import annotations

import csv
import json
import math
import re
import unicodedata
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
ADP_DIR = ROOT / "ADP Rankings"
OUT_DIR = ROOT / "rankings" / "adp-sources"
ESPN_TSV = ROOT / "espn-2026-draft-ranks-ppr-standard-superflex.tsv"
UNRANKED = 999.0
# Kickers/DST before ~round 12 in a 12-team draft are site outliers (FP ESPN ~70, PFF ~85).
K_DST_MIN_CONSENSUS_ADP = 140.0
# PFF stamps this ADP on everyone they don't actually rank. Skip those rows; use RANK for the rest.
PFF_ADP_FLOOR_MIN = 169.5
PFF_ADP_FLOOR_MAX = 171.0
POS_PREF = {"QB": 5, "RB": 5, "WR": 5, "TE": 5, "DST": 3, "DEF": 3, "K": 1, "": 0}

TEAM_CANON = {
    "JAC": "JAX",
    "WSH": "WAS",
    "LA": "LAR",
    "ARZ": "ARI",
    "AZ": "ARI",
    "GNB": "GB",
    "KAN": "KC",
    "SFO": "SF",
    "NWE": "NE",
    "NOR": "NO",
    "TAM": "TB",
    "OAK": "LV",
    "SD": "LAC",
    "SDG": "LAC",
    "STL": "LAR",
    "PHL": "PHI",
    "LVR": "LV",
}

TEAM_ABBRS = {
    "ARI", "ATL", "BAL", "BUF", "CAR", "CHI", "CIN", "CLE", "DAL", "DEN", "DET",
    "GB", "HOU", "IND", "JAX", "JAC", "KC", "LAC", "LAR", "LA", "LV", "LVR", "MIA",
    "MIN", "NE", "NO", "NYG", "NYJ", "PHI", "PIT", "SEA", "SF", "TB", "TEN", "WAS",
    "WSH",
}

FULL_TEAM_TO_ABBR = {
    "arizona cardinals": "ARI",
    "atlanta falcons": "ATL",
    "baltimore ravens": "BAL",
    "buffalo bills": "BUF",
    "carolina panthers": "CAR",
    "chicago bears": "CHI",
    "cincinnati bengals": "CIN",
    "cleveland browns": "CLE",
    "dallas cowboys": "DAL",
    "denver broncos": "DEN",
    "detroit lions": "DET",
    "green bay packers": "GB",
    "houston texans": "HOU",
    "indianapolis colts": "IND",
    "jacksonville jaguars": "JAX",
    "kansas city chiefs": "KC",
    "las vegas raiders": "LV",
    "los angeles chargers": "LAC",
    "los angeles rams": "LAR",
    "miami dolphins": "MIA",
    "minnesota vikings": "MIN",
    "new england patriots": "NE",
    "new orleans saints": "NO",
    "new york giants": "NYG",
    "new york jets": "NYJ",
    "philadelphia eagles": "PHI",
    "pittsburgh steelers": "PIT",
    "san francisco 49ers": "SF",
    "seattle seahawks": "SEA",
    "tampa bay buccaneers": "TB",
    "tennessee titans": "TEN",
    "washington commanders": "WAS",
}

ABBR_TO_FULL = {abbr: name.title() for name, abbr in FULL_TEAM_TO_ABBR.items()}
ABBR_TO_FULL["SF"] = "San Francisco 49ers"
ABBR_TO_FULL["KC"] = "Kansas City Chiefs"
ABBR_TO_FULL["TB"] = "Tampa Bay Buccaneers"
ABBR_TO_FULL["GB"] = "Green Bay Packers"
ABBR_TO_FULL["LV"] = "Las Vegas Raiders"
ABBR_TO_FULL["NE"] = "New England Patriots"
ABBR_TO_FULL["NO"] = "New Orleans Saints"
ABBR_TO_FULL["NYG"] = "New York Giants"
ABBR_TO_FULL["NYJ"] = "New York Jets"
ABBR_TO_FULL["LAC"] = "Los Angeles Chargers"
ABBR_TO_FULL["LAR"] = "Los Angeles Rams"

NICKNAME_TO_ABBR = {
    "cardinals": "ARI", "arizona": "ARI",
    "falcons": "ATL", "atlanta": "ATL",
    "ravens": "BAL", "baltimore": "BAL",
    "bills": "BUF", "buffalo": "BUF",
    "panthers": "CAR", "carolina": "CAR",
    "bears": "CHI", "chicago": "CHI",
    "bengals": "CIN", "cincinnati": "CIN",
    "browns": "CLE", "cleveland": "CLE",
    "cowboys": "DAL", "dallas": "DAL",
    "broncos": "DEN", "denver": "DEN",
    "lions": "DET", "detroit": "DET",
    "packers": "GB", "greenbay": "GB",
    "texans": "HOU", "houston": "HOU",
    "colts": "IND", "indianapolis": "IND",
    "jaguars": "JAX", "jags": "JAX", "jacksonville": "JAX",
    "chiefs": "KC",
    "raiders": "LV",
    "chargers": "LAC",
    "rams": "LAR",
    "dolphins": "MIA", "miami": "MIA",
    "vikings": "MIN", "minnesota": "MIN",
    "patriots": "NE", "pats": "NE",
    "saints": "NO",
    "giants": "NYG",
    "jets": "NYJ",
    "eagles": "PHI", "philadelphia": "PHI",
    "steelers": "PIT", "pittsburgh": "PIT",
    "49ers": "SF", "niners": "SF",
    "seahawks": "SEA", "seattle": "SEA",
    "buccaneers": "TB", "bucs": "TB", "tampa": "TB",
    "titans": "TEN", "tennessee": "TEN",
    "commanders": "WAS", "washington": "WAS",
}

DST_STRIP_RE = re.compile(
    r"[\s,]*(\bD\s*/\s*ST\b|\bDST\b|\bDEF\b|\bD/ST\b)\s*$",
    re.I,
)
STATUS_TAIL_RE = re.compile(
    r"\s+(?:-P|NFI-A|NFI|DNR|NA|PUP|SUS|OUT|\bQ\b|\bO\b|\bIR\b|\bP\b)\s*$",
    re.I,
)

POS_CANON = {
    "QB": "QB", "RB": "RB", "WR": "WR", "TE": "TE", "K": "K", "PK": "K",
    "DEF": "DST", "DST": "DST", "D/ST": "DST", "D": "DST", "FB": "RB",
}

# Pos/team when the source file leaves them blank or glues junk onto the name.
KNOWN_SLOTS = {
    "scott matlock": ("RB", "LAC"),
    "ben vansumeren": ("RB", "BUF"),
    "cj williams": ("WR", "JAX"),
}

NFL_DST_TEAMS = [
    "ARI", "ATL", "BAL", "BUF", "CAR", "CHI", "CIN", "CLE", "DAL", "DEN", "DET",
    "GB", "HOU", "IND", "JAX", "KC", "LAC", "LAR", "LV", "MIA", "MIN", "NE", "NO",
    "NYG", "NYJ", "PHI", "PIT", "SEA", "SF", "TB", "TEN", "WAS",
]

# Copy D/ST ADP from a 1QB board that already has all 32.
DST_FILL_FROM = {
    "ppr_season_superflex": "ppr_season_1qb",
    "half_ppr_season_superflex": "half_ppr_season_1qb",
    "standard_season_superflex": "standard_season_1qb",
    "ppr_dynasty_1qb": "ppr_season_1qb",
    "half_ppr_dynasty_1qb": "half_ppr_season_1qb",
    "standard_dynasty_1qb": "standard_season_1qb",
    "ppr_dynasty_superflex": "ppr_season_superflex",
    "half_ppr_dynasty_superflex": "half_ppr_season_superflex",
    "standard_dynasty_superflex": "standard_season_superflex",
}

SOURCE_LABELS = {
    "espn": "ESPN",
    "cbs": "CBS",
    "yahoo": "Yahoo",
    "sleeper": "Sleeper",
    "fantasypros": "FantasyPros",
    "pff": "PFF",
    "rtsports": "RTSports",
    "fantrax": "Fantrax",
    "underdog": "Underdog",
    "draftsharks": "DraftSharks",
    "nfl": "NFL",
}

# Preferred native copy wins when two extracts of the same site are close.
NATIVE_PRIORITY = {
    "espn": [
        "espn_sf_blend",
        "espn_dynasty_fill",
        "espn_ppr_order",
        "espn_pdf",
        "espn_tsv",
        "espn_sheet",
        "fp_adp",
        "draftsharks",
    ],
    "cbs": ["cbs_sheet", "fp_adp", "draftsharks"],
    "yahoo": ["yahoo_sf_blend", "yahoo_list", "yahoo_sheet", "fp_adp", "draftsharks"],
    "sleeper": ["sleeper_api", "sleeper_tsv", "sleeper_sheet", "fp_adp", "draftsharks"],
}

SUFFIX_RE = re.compile(r"\s+(jr\.?|sr\.?|ii|iii|iv|v)$", re.I)
POS_TOKEN_RE = re.compile(r"^(QB|RB|WR|TE|K|PK|FB|DEF|DST|D/ST|D)-?\d*$", re.I)
# Trailing Q/O status is stripped by STATUS_TAIL_RE. Do not use \bQ\b or O$ here —
# those eat initials ("Q. Judkins" → ". Judkins") and names ending in o (Skattebo, Pacheco).
BYE_RE = re.compile(r"\(Q\)|\bIR\b|\bPUP\b|\bSUS\b", re.I)


class PlayerAdp:
    __slots__ = ("name", "pos", "team", "adp")

    def __init__(self, name: str, pos: str, team: str, adp: float):
        self.name = name.strip()
        self.pos = canon_pos(pos)
        self.team = canon_team(team)
        self.adp = float(adp)
        dst = resolve_defense(self.name, self.pos, self.team)
        if dst:
            self.name, self.pos, self.team = dst
        slot = KNOWN_SLOTS.get(re.sub(r"[^a-z0-9]+", " ", self.name.lower()).strip())
        if slot:
            self.pos = self.pos or slot[0]
            self.team = self.team or slot[1]

    def as_dict(self) -> dict[str, Any]:
        return {"name": self.name, "pos": self.pos, "team": self.team, "adp": round(self.adp, 2)}


def canon_team(raw: str | None) -> str:
    if not raw:
        return ""
    s = unicodedata.normalize("NFKC", str(raw)).strip()
    if not s:
        return ""
    low = s.lower()
    if low in FULL_TEAM_TO_ABBR:
        return FULL_TEAM_TO_ABBR[low]
    u = re.sub(r"[^A-Za-z]", "", s).upper()
    if u in TEAM_CANON:
        return TEAM_CANON[u]
    if u in TEAM_ABBRS:
        return TEAM_CANON.get(u, u)
    return ""


def canon_pos(raw: str | None) -> str:
    if not raw:
        return ""
    s = str(raw).strip().upper().replace(" ", "")
    m = POS_TOKEN_RE.match(s)
    if m:
        token = re.sub(r"-?\d+$", "", m.group(1).upper())
        return POS_CANON.get(token, token)
    token = re.sub(r"[^A-Z/]", "", s)
    token = re.sub(r"-?\d+$", "", token)
    return POS_CANON.get(token, token if token in POS_CANON else "")


def defense_abbr_from_name(raw: str) -> str:
    """Map 'Bears', 'Buffalo Bills DST', 'Broncos D/ST' to a team abbr."""
    s = DST_STRIP_RE.sub("", cell_str(raw)).strip(" ,-/")
    if not s:
        return ""
    low = re.sub(r"\s+", " ", s.lower())
    low = low.replace(".", "")
    if low in FULL_TEAM_TO_ABBR:
        return FULL_TEAM_TO_ABBR[low]
    nick = re.sub(r"[^a-z0-9]", "", low)
    if nick in NICKNAME_TO_ABBR:
        return NICKNAME_TO_ABBR[nick]
    return canon_team(s)


def resolve_defense(name: str, pos: str, team: str) -> tuple[str, str, str] | None:
    """If this row is a D/ST, return (canonical full name, DST, abbr)."""
    pos_c = canon_pos(pos)
    team_c = canon_team(team) or defense_abbr_from_name(name)
    looks_dst = pos_c in {"DST", "DEF"} or bool(DST_STRIP_RE.search(name or ""))
    if not looks_dst:
        # Whole-name nickname only ("Bears", "Texans") — not "Chase Brown".
        tokens = (name or "").split()
        if len(tokens) == 1 and defense_abbr_from_name(name):
            looks_dst = True
            team_c = team_c or defense_abbr_from_name(name)
        elif defense_abbr_from_name(name) and not re.search(r"\b(QB|RB|WR|TE|K)\b", name or "", re.I):
            # "Houston Texans" / "Buffalo Bills DST" with empty pos
            full = DST_STRIP_RE.sub("", name or "").strip()
            if full.lower() in FULL_TEAM_TO_ABBR:
                looks_dst = True
                team_c = team_c or FULL_TEAM_TO_ABBR[full.lower()]
    if not looks_dst or not team_c:
        return None
    full = ABBR_TO_FULL.get(team_c, "")
    if not full:
        return None
    return full, "DST", team_c


def norm_name(raw: str) -> str:
    s = unicodedata.normalize("NFKC", raw or "")
    s = s.replace("\u2019", "'").replace("`", "'").replace("\u2018", "'")
    s = s.replace(".", " ")
    s = re.sub(r"['\u00B4]", "", s)
    s = re.sub(r"[^A-Za-z0-9\s\-]", " ", s)
    s = re.sub(r"\s+", " ", s).strip().lower()
    while True:
        n = SUFFIX_RE.sub("", s).strip()
        if n == s:
            break
        s = n
    return s


def parse_adp(value: Any, *, cap: float | None = UNRANKED) -> float | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        n = float(value)
        if n <= 0:
            return None
        if cap is not None and n >= cap:
            return None
        return n
    s = str(value).strip()
    if not s or s in {"-", "—", "–", "�", "N/A", "NA", "null"}:
        return None
    s = s.replace("�", "").strip()
    m = re.match(r"^(\d+(?:\.\d+)?)", s)
    if not m:
        return None
    n = float(m.group(1))
    if n <= 0:
        return None
    if cap is not None and n >= cap:
        return None
    return n


def cell_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return ""
    if isinstance(v, float) and math.isnan(v):
        return ""
    return str(v).replace("\xa0", " ").strip()


def parse_player_blob(raw: str) -> tuple[str, str, str]:
    """Return (name, pos, team) from messy site strings."""
    text = cell_str(raw).replace("\r", "\n")
    text = text.replace("\n", " ")
    text = re.sub(r"\s+", " ", text).strip()
    if not text:
        return "", "", ""

    team = ""
    pos = ""

    # Sleeper glues unsigned/rookie pos-rank onto the name: "Tyreek Hill UNSWR82", "Eric McAlister RKWR32"
    glued_rank = re.search(
        r"(?:^|\s)(?:UNS|RK)?(QB|RB|WR|TE|K|PK|DST|DEF|D/ST)\d+\s*$", text, re.I
    )
    if glued_rank:
        pos = pos or canon_pos(glued_rank.group(1))
        text = text[: glued_rank.start()].strip()

    # Free-agent markers: "Tyreek Hill (FA)", "T. HillFA"
    text = re.sub(r"\s*\(\s*FA\s*\)\s*$", "", text, flags=re.I).strip()
    text = re.sub(r"(?<=[A-Za-z])FA$", "", text).strip()

    # Yahoo: "DeVonta Smith\nPhi - WR\n\nQ" → "DeVonta Smith Phi - WR Q"
    for _ in range(4):
        nxt = STATUS_TAIL_RE.sub("", text).strip()
        if nxt == text:
            break
        text = nxt

    # "Jahmyr GibbsDET (6)" / "Jahmyr Gibbs (DET)"
    bye = re.search(r"\((\d{1,2}|null)\)\s*$", text, re.I)
    if bye:
        text = text[: bye.start()].strip()

    paren_team = re.search(r"\(([A-Za-z]{2,3})\)\s*$", text)
    if paren_team and canon_team(paren_team.group(1)):
        team = canon_team(paren_team.group(1))
        text = text[: paren_team.start()].strip()

    # "J.K. Dobbins" must not yield position K from the middle initial.
    text = re.sub(r"\b([A-Z])\.\s*([A-Z])\.(?=\s|$)", r"\1\2.", text)

    # "Det - RB" / "DET - RB" / "Phi - WR Q" (status already stripped)
    dash_pos = re.search(r"[-–]\s*(QB|RB|WR|TE|K|PK|DEF|DST)\b", text, re.I)
    if dash_pos:
        pos = canon_pos(dash_pos.group(1))
        text = text[: dash_pos.start()].strip()

    # glued team on the last token: "Jahmyr GibbsDET", "J. GibbsDET", "A. St. BrownDET"
    # glued team: prefer a known 3-letter then 2-letter suffix ("Walker IIIKC", "GibbsDET")
    if not team:
        for n in (3, 2):
            if len(text) <= n:
                continue
            suffix = text[-n:]
            if suffix.isupper() and canon_team(suffix):
                leftover = text[:-n].strip()
                if leftover and leftover.lower() not in {"jr", "sr"}:
                    team = canon_team(suffix)
                    text = leftover
                    break

    trailing_team = re.search(r"\b([A-Za-z]{2,3})\s*$", text)
    if trailing_team and canon_team(trailing_team.group(1)):
        team = team or canon_team(trailing_team.group(1))
        text = text[: trailing_team.start()].strip()

    # Position only as a leading or trailing token — never a mid-name "K".
    lead_pos = re.match(r"^(QB|RB|WR|TE|PK|K|DEF|DST)\b\s+", text, re.I)
    if lead_pos and not pos:
        pos = canon_pos(lead_pos.group(1))
        text = text[lead_pos.end() :].strip()
    trail_pos = re.search(r"\b(QB|RB|WR|TE|PK|K|DEF|DST)\s*$", text, re.I)
    if trail_pos and not pos:
        pos = canon_pos(trail_pos.group(1))
        text = text[: trail_pos.start()].strip()

    # sleeper "LARWR1" leftover
    glued_pt = re.search(r"([A-Z]{2,3})(QB|RB|WR|TE|K|DST|DEF)\d*$", text)
    if glued_pt and canon_team(glued_pt.group(1)):
        team = team or canon_team(glued_pt.group(1))
        pos = pos or canon_pos(glued_pt.group(2))
        text = text[: glued_pt.start()].strip()

    text = BYE_RE.sub("", text)
    text = re.sub(r"\b([A-Z])\.\s*\.\s+", r"\1. ", text)
    text = re.sub(r"\s+", " ", text).strip(" -,\t")
    parts = text.split()
    if len(parts) >= 2 and canon_team(parts[-1]):
        team = team or canon_team(parts[-1])
        text = " ".join(parts[:-1])
    text = re.sub(r"\s*\(\s*FA\s*\)\s*$", "", text, flags=re.I).strip()
    text = re.sub(r"(?<=[A-Za-z])FA$", "", text).strip()
    text = re.sub(r"\s+FA$", "", text, flags=re.I).strip()
    # "J. Gibbs" keep
    return text, pos, team


def sheet_rows(path: Path, sheet: str) -> list[list[Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet]
        return [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def header_map(row: list[Any]) -> dict[str, int]:
    out: dict[str, int] = {}
    for i, v in enumerate(row):
        key = re.sub(r"\s+", " ", cell_str(v)).strip().lower()
        key = key.replace("\n", " ")
        if key:
            out[key] = i
    return out


def find_header_idx(headers: dict[str, int], *cands: str) -> int | None:
    for c in cands:
        if c in headers:
            return headers[c]
    for key, idx in headers.items():
        for c in cands:
            if c in key:
                return idx
    return None


def is_tier_row(row: list[Any]) -> bool:
    first = cell_str(row[0] if row else "")
    return bool(re.match(r"^tier\s+\d+", first, re.I)) or first.lower() in {"customize tiers"}


def parse_fp_adp(path: Path, sheet: str) -> dict[str, list[PlayerAdp]]:
    rows = sheet_rows(path, sheet)
    if not rows:
        return {}
    headers = header_map(rows[0])
    name_i = find_header_idx(headers, "player (bye)", "player name", "player")
    pos_i = find_header_idx(headers, "pos")
    rank_i = find_header_idx(headers, "rank", "rk")
    if name_i is None:
        return {}

    site_cols: dict[str, int] = {}
    for key, idx in headers.items():
        k = key.replace(" ", "")
        if k in {"espn"}:
            site_cols["espn"] = idx
        elif k in {"cbs"}:
            site_cols["cbs"] = idx
        elif k in {"sleeper"}:
            site_cols["sleeper"] = idx
        elif k in {"yahoo"}:
            site_cols["yahoo"] = idx
        elif k in {"rtsports", "rt sports"}:
            site_cols["rtsports"] = idx
        elif k in {"fantrax"}:
            site_cols["fantrax"] = idx
        elif k in {"nfl"}:
            site_cols["nfl"] = idx

    out: dict[str, list[PlayerAdp]] = {src: [] for src in site_cols}
    out["fantasypros"] = []

    for row in rows[1:]:
        if not row or is_tier_row(row):
            continue
        name, pos, team = parse_player_blob(cell_str(row[name_i] if name_i < len(row) else ""))
        if pos_i is not None and pos_i < len(row):
            col_pos = canon_pos(cell_str(row[pos_i]))
            if col_pos:
                pos = col_pos
        if not name:
            continue
        fp_adp = parse_adp(row[rank_i] if rank_i is not None and rank_i < len(row) else None)
        if fp_adp is not None:
            out["fantasypros"].append(PlayerAdp(name, pos, team, fp_adp))
        for src, idx in site_cols.items():
            if idx >= len(row):
                continue
            adp = parse_adp(row[idx])
            if adp is None:
                continue
            out[src].append(PlayerAdp(name, pos, team, adp))
    return {k: v for k, v in out.items() if v}


def parse_fp_rank_list(path: Path, sheet: str, source: str = "fantasypros") -> dict[str, list[PlayerAdp]]:
    rows = sheet_rows(path, sheet)
    if not rows:
        return {}
    headers = header_map(rows[0])
    name_i = find_header_idx(headers, "player name", "player (bye)", "player")
    pos_i = find_header_idx(headers, "pos")
    rank_i = find_header_idx(headers, "rk", "rank")
    avg_i = find_header_idx(headers, "avg")
    if name_i is None or rank_i is None:
        return {}
    players: list[PlayerAdp] = []
    for row in rows[1:]:
        if not row or is_tier_row(row):
            continue
        name, pos, team = parse_player_blob(cell_str(row[name_i] if name_i < len(row) else ""))
        if pos_i is not None and pos_i < len(row):
            col_pos = canon_pos(cell_str(row[pos_i]))
            if col_pos:
                pos = col_pos
        adp = parse_adp(row[avg_i] if avg_i is not None and avg_i < len(row) else None)
        if adp is None:
            adp = parse_adp(row[rank_i] if rank_i < len(row) else None)
        if not name or adp is None:
            continue
        players.append(PlayerAdp(name, pos, team, adp))
    return {source: players} if players else {}


def parse_pff(path: Path, sheet: str, *, rookie_rank: bool = False) -> dict[str, list[PlayerAdp]]:
    rows = sheet_rows(path, sheet)
    cells: list[str] = []
    for row in rows:
        for v in row:
            s = cell_str(v)
            if s:
                cells.append(s)

    players: list[PlayerAdp] = []
    i = 0
    while i < len(cells) - 2:
        name = cells[i]
        team = canon_team(cells[i + 1])
        pos = canon_pos(cells[i + 2])
        looks_name = bool(re.match(r"^[A-Z][A-Za-z'\.\-]+(?:\s+[A-Z][A-Za-z'\-]+)+$", name)) or bool(
            re.match(r"^[A-Z]\.\s+[A-Z][A-Za-z'\-]+", name)
        )
        if looks_name and team and pos:
            found_adp_label = False
            adp = None
            for j in range(i + 3, min(i + 10, len(cells))):
                if cells[j].upper() == "ADP":
                    found_adp_label = True
                    adp = parse_adp(cells[j - 1]) if j - 1 >= 0 else None
                    break
            rank = None
            for j in range(i - 1, max(-1, i - 6), -1):
                r = parse_adp(cells[j])
                if r is not None and r == int(r) and r < 800:
                    rank = r
                    break
            if rookie_rank:
                use = rank
            else:
                floor = adp is not None and PFF_ADP_FLOOR_MIN <= adp <= PFF_ADP_FLOOR_MAX
                if found_adp_label and (adp is None or floor):
                    i += 3
                    continue
                use = rank if rank is not None else adp
            if use is not None:
                players.append(PlayerAdp(name, pos, team, use))
            i += 3
            continue
        i += 1
    return {"pff": players} if players else {}


def parse_cbs(path: Path, sheet: str) -> dict[str, list[PlayerAdp]]:
    rows = sheet_rows(path, sheet)
    if not rows:
        return {}
    headers = header_map(rows[0])
    name_i = find_header_idx(headers, "player")
    rank_i = find_header_idx(headers, "rank")
    avg_i = find_header_idx(headers, "avg pos", "avg")
    if name_i is None:
        return {}
    players: list[PlayerAdp] = []
    for row in rows[1:]:
        if not row:
            continue
        name, pos, team = parse_player_blob(cell_str(row[name_i] if name_i < len(row) else ""))
        adp = parse_adp(row[avg_i] if avg_i is not None and avg_i < len(row) else None)
        if adp is None:
            adp = parse_adp(row[rank_i] if rank_i is not None and rank_i < len(row) else None)
        if not name or adp is None:
            continue
        players.append(PlayerAdp(name, pos, team, adp))
    return {"cbs": players} if players else {}


def parse_yahoo(path: Path, sheet: str) -> dict[str, list[PlayerAdp]]:
    rows = sheet_rows(path, sheet)
    if not rows:
        return {}
    headers = header_map(rows[0])
    name_i = find_header_idx(headers, "player")
    rank_i = find_header_idx(headers, "rank")
    avg_i = find_header_idx(headers, "avg", "all drafts", "preseason")
    all_i = find_header_idx(headers, "all drafts")
    if name_i is None:
        return {}
    players: list[PlayerAdp] = []
    for row in rows[1:]:
        if not row:
            continue
        name, pos, team = parse_player_blob(cell_str(row[name_i] if name_i < len(row) else ""))
        adp = None
        if all_i is not None and all_i < len(row):
            adp = parse_adp(row[all_i])
        if adp is None and avg_i is not None and avg_i < len(row):
            adp = parse_adp(row[avg_i])
        if adp is None and rank_i is not None and rank_i < len(row):
            adp = parse_adp(row[rank_i])
        if not name or adp is None:
            continue
        players.append(PlayerAdp(name, pos, team, adp))
    return {"yahoo": players} if players else {}


def parse_numbered_name_list(path: Path) -> list[PlayerAdp]:
    """Plain '1. Name' lists (Yahoo Superflex paste)."""
    if not path.exists():
        return []
    players: list[PlayerAdp] = []
    for line in path.read_text(encoding="utf-8").splitlines():
        text = line.strip()
        if not text or text.startswith("#"):
            continue
        m = re.match(r"^(\d+)\.\s+(.+)$", text)
        if not m:
            continue
        rank = float(m.group(1))
        name, pos, team = parse_player_blob(m.group(2).strip())
        dst = resolve_defense(name, pos, team)
        if dst:
            name, pos, team = dst
        if not name:
            continue
        players.append(PlayerAdp(name, pos, team, rank))
    return players


def parse_espn_labeled_list(path: Path) -> list[PlayerAdp]:
    """ESPN paste: '1. Name, LAR -- WR1 (Age: 25-3)' or '1. Name, Bills (QB1)'."""
    if not path.exists():
        return []
    players: list[PlayerAdp] = []
    for line in path.read_text(encoding="utf-8").splitlines():
        text = line.strip().rstrip(":").strip()
        if not text or text.startswith("#"):
            continue
        m = re.match(r"^(\d+)\.\s+(.+)$", text)
        if not m:
            continue
        rank = float(m.group(1))
        rest = m.group(2).strip()
        rest = re.sub(r"\s*\(Age:[^)]*\)\s*$", "", rest, flags=re.I).strip()
        pos = ""
        pos_paren = re.search(r"\(([A-Za-z]{1,3})\d+\)\s*$", rest)
        if pos_paren:
            pos = canon_pos(pos_paren.group(1))
            rest = rest[: pos_paren.start()].strip()
        pos_dash = re.search(r"--\s*([A-Za-z]{1,3})\d+\s*$", rest)
        if pos_dash:
            pos = pos or canon_pos(pos_dash.group(1))
            rest = rest[: pos_dash.start()].strip()
        name, team = rest, ""
        if "," in rest:
            name, team_raw = rest.rsplit(",", 1)
            name = name.strip()
            team_raw = team_raw.strip()
            team = canon_team(team_raw) or defense_abbr_from_name(team_raw)
        else:
            name, blob_pos, team = parse_player_blob(rest)
            pos = pos or blob_pos
        dst = resolve_defense(name, pos, team)
        if dst:
            name, pos, team = dst
        if not name:
            continue
        players.append(PlayerAdp(name, pos, team, rank))
    return players


def parse_draftsharks(path: Path, sheet: str) -> dict[str, list[PlayerAdp]]:
    """Multi-site DraftSharks export: row1 site names, row2 column types."""
    rows = sheet_rows(path, sheet)
    if len(rows) < 3:
        return {}
    site_row = [cell_str(v) for v in rows[0]]
    type_row = [re.sub(r"\s+", " ", cell_str(v)).strip().lower() for v in rows[1]]

    # Fill forward site names across merged header cells.
    filled: list[str] = []
    last = ""
    for v in site_row:
        if v:
            last = v
        filled.append(last)

    def site_key(label: str) -> str | None:
        k = label.strip().lower()
        if k in {"cbs"}:
            return "cbs"
        if k in {"espn"}:
            return "espn"
        if k in {"sleeper"}:
            return "sleeper"
        if k in {"yahoo"}:
            return "yahoo"
        # Consensus ADP is the usable DraftSharks number. DS Rank is an internal id (981, 765…).
        if k in {"consensus"}:
            return "draftsharks"
        return None

    col_map: dict[str, int] = {}
    name_i = None
    for i, typ in enumerate(type_row):
        site = site_key(filled[i] if i < len(filled) else "")
        if "player" in typ:
            name_i = i
        if site and typ == "adp":
            col_map[site] = i

    if name_i is None:
        return {}

    out: dict[str, list[PlayerAdp]] = {k: [] for k in col_map}
    for row in rows[2:]:
        if not row:
            continue
        name, pos, team = parse_player_blob(cell_str(row[name_i] if name_i < len(row) else ""))
        if not name:
            continue
        for src, idx in col_map.items():
            if idx >= len(row):
                continue
            adp = parse_adp(row[idx])
            if adp is None:
                continue
            out[src].append(PlayerAdp(name, pos, team, adp))
    return {k: v for k, v in out.items() if v}


def parse_sleeper_ds(path: Path, sheet: str, source: str = "sleeper", *, prefer_rank: bool = False) -> dict[str, list[PlayerAdp]]:
    rows = sheet_rows(path, sheet)
    if len(rows) < 2:
        return {}
    # header may be on row 0 or 1
    header_idx = 0
    for i, row in enumerate(rows[:3]):
        joined = " ".join(cell_str(c).lower() for c in row)
        if "player" in joined and ("adp" in joined or "rank" in joined):
            header_idx = i
            break
    headers = header_map(rows[header_idx])
    name_i = find_header_idx(headers, "player")
    adp_i = find_header_idx(headers, "adp")
    rank_i = find_header_idx(headers, "ds rank", "rank")
    if name_i is None:
        return {}
    # If this sheet also has a Sleeper-labeled ADP to the right, prefer that later via draftsharks parser.
    players: list[PlayerAdp] = []
    for row in rows[header_idx + 1 :]:
        if not row:
            continue
        name, pos, team = parse_player_blob(cell_str(row[name_i] if name_i < len(row) else ""))
        rank_v = parse_adp(row[rank_i] if rank_i is not None and rank_i < len(row) else None)
        adp_v = parse_adp(row[adp_i] if adp_i is not None and adp_i < len(row) else None)
        adp = (rank_v if prefer_rank else adp_v) or (adp_v if prefer_rank else rank_v)
        if not name or adp is None:
            continue
        players.append(PlayerAdp(name, pos, team, adp))
    return {source: players} if players else {}


def parse_underdog(path: Path) -> list[PlayerAdp]:
    players: list[PlayerAdp] = []
    with path.open(encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            first = (row.get("firstName") or "").strip()
            last = (row.get("lastName") or "").strip()
            name = f"{first} {last}".strip()
            adp = parse_adp(row.get("adp"))
            pos = canon_pos(row.get("slotName") or "")
            team = canon_team(row.get("teamName") or "")
            if not name or adp is None:
                continue
            players.append(PlayerAdp(name, pos, team, adp))
    return players


def parse_sleeper_tsv(path: Path) -> list[PlayerAdp]:
    players: list[PlayerAdp] = []
    with path.open(encoding="utf-8") as f:
        reader = csv.DictReader(f, delimiter="\t")
        for row in reader:
            name, blob_pos, blob_team = parse_player_blob(row.get("Player") or "")
            adp = parse_adp(row.get("Rank"))
            pos = canon_pos(row.get("Position") or "") or blob_pos
            team = canon_team(row.get("Team") or "") or blob_team
            if not name or adp is None:
                continue
            players.append(PlayerAdp(name, pos, team, adp))
    return players


def parse_espn_tsv(path: Path, *, cap: float | None = UNRANKED) -> dict[str, list[PlayerAdp]]:
    out: dict[str, list[PlayerAdp]] = {"ppr": [], "standard": [], "superflex": []}
    with path.open(encoding="utf-8") as f:
        reader = csv.DictReader(f, delimiter="\t")
        for row in reader:
            name, _, _ = parse_player_blob(row.get("Player") or "")
            if not name:
                continue
            for key, bucket in (("PPR", "ppr"), ("Standard", "standard"), ("Superflex", "superflex")):
                adp = parse_adp(row.get(key), cap=cap)
                if adp is None:
                    continue
                out[bucket].append(PlayerAdp(name, "", "", adp))
    return {k: v for k, v in out.items() if v}


ESPN_PDF_ENTRY_RE = re.compile(
    r"(\d{1,3})\.\s+\((QB|RB|WR|TE|K|DST|D/ST|DEF)\d+\)\s+"
    r"(.+?)"
    r"(?=\s+\d{1,3}\.\s+\(|$)",
    re.S,
)
ESPN_PDF_TEAM_RE = re.compile(r"^(.+?),\s*([A-Z]{2,3}|FA)\b")
ESPN_PDF_ROOKIE_RE = re.compile(
    r"^([A-Za-z][A-Za-z.'\-]+(?:\s+[A-Za-z][A-Za-z.'\-]+)*)\s+\d{4}-"
)


def parse_espn_pdf(path: Path) -> dict[str, list[PlayerAdp]]:
    """ESPN cheat sheets: 4 columns of `1. (RB1) Jahmyr Gibbs, DET $57 6`.

    Dynasty sheets put a rookie board in the last column (`1. (RB1) Jeremiyah Love 2026-1`).
    """
    empty: dict[str, list[PlayerAdp]] = {"overall": [], "rookies": []}
    if not path.exists():
        return empty
    try:
        from pypdf import PdfReader
    except ImportError:
        return empty
    text = " ".join((page.extract_text() or "") for page in PdfReader(str(path)).pages)
    text = re.sub(r"\s+", " ", text).strip()
    overall: dict[int, PlayerAdp] = {}
    rookies: dict[int, PlayerAdp] = {}
    for m in ESPN_PDF_ENTRY_RE.finditer(text):
        rank = int(m.group(1))
        pos = m.group(2)
        blob = m.group(3).strip()
        blob = re.sub(r"\s+(Salary Cap Value|Bye week)\b.*$", "", blob, flags=re.I).strip()
        team_m = ESPN_PDF_TEAM_RE.match(blob)
        if team_m:
            name = team_m.group(1).strip(" ,")
            team = team_m.group(2)
            if name and rank not in overall:
                overall[rank] = PlayerAdp(name, pos, team, float(rank))
            continue
        rook_m = ESPN_PDF_ROOKIE_RE.match(blob)
        if rook_m:
            name = rook_m.group(1).strip(" ,")
            if name and rank not in rookies:
                rookies[rank] = PlayerAdp(name, pos, "", float(rank))
    return {
        "overall": [overall[k] for k in sorted(overall)],
        "rookies": [rookies[k] for k in sorted(rookies)],
    }


def load_rookie_names_from_files() -> list[str]:
    names: list[str] = []
    files = [
        ROOT / "rankings" / "generated" / "rookies_dynasty_1qb.csv",
        ROOT / "rankings" / "generated" / "rookies_dynasty_superflex.csv",
        ROOT / "data" / "rookies_2026_board.csv",
    ]
    for path in files:
        if not path.exists():
            continue
        with path.open(encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                name = (row.get("name") or row.get("Player") or "").strip()
                if name:
                    names.append(name)
    return names


def names_are_same_player(a: str, b: str) -> bool:
    if name_parts(a)[0] == name_parts(b)[0]:
        return True
    return first_names_compatible(a, b, teams_agree=False) and last_names_compatible(
        a, b, teams_agree=False
    )


def merge_espn_overall(*lists: list[PlayerAdp]) -> list[PlayerAdp]:
    """Keep the best (lowest) ESPN rank per name so the PDF 300 overwrites TSV on overlap."""
    by_name: dict[str, PlayerAdp] = {}
    for lst in lists:
        for p in lst:
            n = name_parts(p.name)[0]
            if not n:
                continue
            prev = by_name.get(n)
            if prev is None or p.adp < prev.adp:
                by_name[n] = p
    return sorted(by_name.values(), key=lambda p: (p.adp, p.name))


def derive_sf_espn_from_1qb(
    one_qb: list[PlayerAdp],
    two_qb: list[PlayerAdp],
) -> list[PlayerAdp]:
    """1QB skill order, with QBs inserted at their Superflex overall ranks.

    Non-QBs keep their 1QB relative order. Each Superflex QB is inserted at that
    QB's overall Superflex rank (processed 1, then 2, then 4, …): whoever sat in
    that slot shifts down one.
    """
    def is_qb(p: PlayerAdp) -> bool:
        if (p.pos or "").upper() == "QB":
            return True
        return any(
            names_are_same_player(p.name, o.name) and (o.pos or "").upper() == "QB"
            for o in one_qb
        )

    def already_in(pool: list[PlayerAdp], name: str) -> bool:
        return any(names_are_same_player(p.name, name) for p in pool)

    def match_1qb(qb: PlayerAdp) -> PlayerAdp:
        for p in one_qb:
            if names_are_same_player(p.name, qb.name):
                return p
        return qb

    one_sorted = sorted(one_qb, key=lambda p: (p.adp, p.name))
    two_sorted = sorted(two_qb, key=lambda p: (p.adp, p.name))
    two_qbs = [p for p in two_sorted if is_qb(p)]

    def is_sf_qb(p: PlayerAdp) -> bool:
        if is_qb(p):
            return True
        return any(names_are_same_player(p.name, q.name) for q in two_qbs)

    out = [p for p in one_sorted if not is_sf_qb(p)]
    for rank, qb in enumerate(two_sorted, 1):
        if not is_qb(qb):
            continue
        if already_in(out, qb.name):
            continue
        row = match_1qb(qb)
        idx = min(max(rank - 1, 0), len(out))
        out.insert(idx, row)
    for p in one_sorted:
        if is_sf_qb(p) and not already_in(out, p.name):
            out.append(p)
    return [
        PlayerAdp(p.name, p.pos, p.team, float(i))
        for i, p in enumerate(out, 1)
    ]


def overlay_identity(target: list[PlayerAdp], *sources: list[PlayerAdp]) -> list[PlayerAdp]:
    """Fill missing pos/team on `target` from earlier Yahoo lists."""
    catalog = [p for src in sources for p in src]
    out: list[PlayerAdp] = []
    for p in target:
        pos, team = p.pos, p.team
        if not pos or not team:
            for other in catalog:
                if names_are_same_player(p.name, other.name):
                    pos = pos or other.pos
                    team = team or other.team
                    break
        out.append(PlayerAdp(p.name, pos, team, p.adp))
    return out


def derive_rookie_ranks(overall: list[PlayerAdp], rookie_names: list[str]) -> list[PlayerAdp]:
    """Walk a redraft board and assign dense 1..n ranks to 2026 rookies in that order."""
    known = [n for n in rookie_names if n.strip()]
    if not known:
        return []
    out: list[PlayerAdp] = []
    seen: set[str] = set()
    for p in sorted(overall, key=lambda x: (x.adp, x.name)):
        n = name_parts(p.name)[0]
        if not n or n in seen:
            continue
        if not any(names_are_same_player(p.name, r) for r in known):
            continue
        seen.add(n)
        out.append(PlayerAdp(p.name, p.pos, p.team, float(len(out) + 1)))
    return out


def fill_espn_from_redraft(
    dynasty_ranked: list[PlayerAdp],
    redraft: list[PlayerAdp],
    universe: list[str],
) -> list[PlayerAdp]:
    """Keep ESPN dynasty ranks, then append missing board players in redraft order."""
    out = list(dynasty_ranked)
    seen = {name_parts(p.name)[0] for p in out}
    ranked_names = [p.name for p in out]
    uni_norms = {name_parts(u)[0] for u in universe if u.strip()}
    uni_names = [u for u in universe if u.strip()]

    def is_seen(name: str) -> bool:
        n = name_parts(name)[0]
        if not n:
            return True
        if n in seen:
            return True
        if any(names_are_same_player(name, r) for r in ranked_names):
            seen.add(n)
            return True
        return False

    def wanted(name: str) -> bool:
        n = name_parts(name)[0]
        if n in uni_norms:
            return True
        for u in uni_names:
            if is_attachable_short(u) or is_attachable_short(name):
                continue
            if names_are_same_player(name, u):
                return True
        return False

    next_rank = int(max((p.adp for p in out), default=0)) + 1 if out else 1
    for p in sorted(redraft, key=lambda x: (x.adp, x.name)):
        if is_seen(p.name) or not wanted(p.name):
            continue
        n = name_parts(p.name)[0]
        seen.add(n)
        ranked_names.append(p.name)
        out.append(PlayerAdp(p.name, p.pos, p.team, float(next_rank)))
        next_rank += 1
    return out


def append_missing_rookies(primary: list[PlayerAdp], extra: list[PlayerAdp]) -> list[PlayerAdp]:
    have = {name_parts(p.name)[0] for p in primary}
    out = list(primary)
    for p in extra:
        n = name_parts(p.name)[0]
        if not n or n in have:
            continue
        have.add(n)
        out.append(PlayerAdp(p.name, p.pos, p.team, float(len(out) + 1)))
    return out


def name_parts(name: str) -> tuple[str, str, str]:
    n = norm_name(name)
    parts = n.split()
    if not parts:
        return "", "", ""
    return n, parts[0], parts[-1]


def is_short_first(name: str) -> bool:
    _, first, _ = name_parts(name)
    return len(first) <= 1


def is_last_name_only(name: str) -> bool:
    s = name.strip()
    if re.match(r"^[\.']+\s+\S+", s):
        return True
    n, _, last = name_parts(s)
    parts = n.split()
    return len(parts) == 1 and len(last) >= 4


def is_attachable_short(name: str) -> bool:
    return is_short_first(name) or is_last_name_only(name)


def last_name_truncation(la: str, lb: str) -> bool:
    """skatteb/skattebo, pachec/pacheco — prefix, extra 1–2 letters, short last >= 5."""
    if not la or not lb or la == lb:
        return False
    short, long_ = (la, lb) if len(la) <= len(lb) else (lb, la)
    if len(short) < 5:
        return False
    extra = len(long_) - len(short)
    if extra < 1 or extra > 2:
        return False
    return long_.startswith(short)


def last_names_compatible(a: str, b: str, *, teams_agree: bool) -> bool:
    _, _, la = name_parts(a)
    _, _, lb = name_parts(b)
    if not la or not lb:
        return False
    if la == lb:
        return True
    return teams_agree and last_name_truncation(la, lb)


FIRST_NAME_ALIASES = {
    "hollywood": "marquise",
    "cam": "cameron",
    "chig": "chigoziem",
    "kenny": "kenneth",
    "tank": "nathaniel",
    "andy": "andres",
}
_REVERSE_FIRST = defaultdict(set)
for _short, _full in FIRST_NAME_ALIASES.items():
    _REVERSE_FIRST[_full].add(_short)


def first_name_variants(first: str) -> set[str]:
    out = {first}
    if first in FIRST_NAME_ALIASES:
        out.add(FIRST_NAME_ALIASES[first])
    out.update(_REVERSE_FIRST.get(first, ()))
    return {x for x in out if x}


def slot_full_firsts(players: list[PlayerAdp]) -> dict[tuple[str, str, str], set[str]]:
    """Full first names keyed by (last, pos, team) so B. Robinson can see Bijan vs Brian."""
    out: dict[tuple[str, str, str], set[str]] = defaultdict(set)
    for p in players:
        _, first, last = name_parts(p.name)
        if not last or len(first) <= 1:
            continue
        if p.pos and p.team:
            out[last, p.pos, p.team].add(first)
        if p.pos:
            out[last, p.pos, ""].add(first)
        out[last, "", ""].add(first)
    return out


def generation_suffix(name: str) -> str:
    """jr/sr/ii from the display name, before norm_name strips them.

    Yahoo lists Bijan as 'B. Robinson' and Brian as 'B. Robinson Jr.' — same
    initial+last after Jr is stripped, so matching has to keep the suffix.
    """
    s = unicodedata.normalize("NFKC", name or "")
    s = s.replace("\u2019", "'").replace(".", " ")
    s = re.sub(r"[^A-Za-z0-9\s\-]", " ", s)
    s = re.sub(r"\s+", " ", s).strip().lower()
    m = re.search(r"\s+(jr|sr|ii|iii|iv|v)$", s)
    if not m:
        return ""
    token = m.group(1)
    return "jr" if token == "jr" else token


def prefer_generation(target: PlayerAdp, hits: list[PlayerAdp]) -> list[PlayerAdp]:
    if len(hits) <= 1:
        return hits
    want = generation_suffix(target.name)
    if want:
        matched = [h for h in hits if generation_suffix(h.name) == want]
        return matched or hits
    without = [h for h in hits if not generation_suffix(h.name)]
    return without or hits


def initial_collision(
    a: str,
    b: str,
    pos: str,
    team: str,
    slot_firsts: dict[tuple[str, str, str], set[str]],
) -> bool:
    """True when a 1-letter first name could mean two people (B. Robinson → Bijan and Brian)."""
    _, fa, la = name_parts(a)
    _, fb, lb = name_parts(b)
    last = la if la == lb else (la or lb)
    initial = ""
    if len(fa) == 1 and len(fb) > 1:
        initial = fa
    elif len(fb) == 1 and len(fa) > 1:
        initial = fb
    else:
        return False
    firsts: set[str] = set()
    if last and pos and team:
        firsts |= slot_firsts.get((last, pos, team), set())
    if last and pos:
        firsts |= slot_firsts.get((last, pos, ""), set())
    if last:
        firsts |= slot_firsts.get((last, "", ""), set())
    hits = {f for f in firsts if f.startswith(initial)}
    return len(hits) > 1


def first_names_compatible(a: str, b: str, *, teams_agree: bool = False) -> bool:
    if is_last_name_only(a) or is_last_name_only(b):
        return True
    _, fa, _ = name_parts(a)
    _, fb, _ = name_parts(b)
    if not fa or not fb:
        return False
    for xa in first_name_variants(fa):
        for xb in first_name_variants(fb):
            if xa == xb:
                return True
            if len(xa) == 1 and xb.startswith(xa):
                if len(xb) <= 2:
                    if teams_agree:
                        return True
                    continue
                return True
            if len(xb) == 1 and xa.startswith(xb):
                if len(xa) <= 2:
                    if teams_agree:
                        return True
                    continue
                return True
            if min(len(xa), len(xb)) >= 3 and (xa.startswith(xb) or xb.startswith(xa)):
                return True
    return False


def slots_compatible(a_pos: str, b_pos: str, a_team: str, b_team: str) -> bool:
    if a_pos and b_pos and a_pos != b_pos:
        return False
    if a_team and b_team and a_team != b_team:
        return False
    return True


def identity_keys(p: PlayerAdp, *, fuzzy: bool = False) -> list[tuple[str, str, str]]:
    n, first, last = name_parts(p.name)
    pos = p.pos
    team = p.team
    if pos == "DST" and team:
        return [("__dst__", "DST", team), (n, "DST", team), (n, "DST", "")]
    keys = [
        (n, pos, team),
        (n, pos, ""),
        (n, "", team),
        (n, "", ""),
    ]
    if fuzzy and first and last:
        keys.append((f"{first[:1]} {last}", pos, team))
        keys.append((f"{first[:1]} {last}", pos, ""))
        if last and pos:
            keys.append((f"last:{last}", pos, team))
            keys.append((f"last:{last}", pos, ""))
    return keys


def index_players(players: list[PlayerAdp]) -> dict[tuple[str, str, str], list[PlayerAdp]]:
    idx: dict[tuple[str, str, str], list[PlayerAdp]] = defaultdict(list)
    for p in players:
        seen: set[tuple[str, str, str]] = set()
        for k in identity_keys(p, fuzzy=True):
            if k in seen:
                continue
            seen.add(k)
            idx[k].append(p)
    return idx


def match_player(
    target: PlayerAdp,
    idx: dict[tuple[str, str, str], list[PlayerAdp]],
    *,
    slot_firsts: dict[tuple[str, str, str], set[str]] | None = None,
) -> PlayerAdp | None:
    for k in identity_keys(target, fuzzy=True):
        hits = idx.get(k) or []
        compatible: list[PlayerAdp] = []
        for h in hits:
            if not slots_compatible(target.pos, h.pos, target.team, h.team):
                continue
            teams_agree = bool(target.team and h.team and target.team == h.team)
            if not last_names_compatible(target.name, h.name, teams_agree=teams_agree):
                continue
            # Initials must match the full first name (J. Allen → Josh, not Kyle).
            if not first_names_compatible(target.name, h.name, teams_agree=teams_agree):
                continue
            compatible.append(h)
        compatible = prefer_generation(target, compatible)
        if slot_firsts:
            kept = [
                h
                for h in compatible
                if not initial_collision(
                    target.name, h.name, target.pos or h.pos, target.team or h.team, slot_firsts
                )
                or generation_suffix(target.name)
                or generation_suffix(h.name)
            ]
            if kept:
                compatible = kept
        if len(compatible) == 1:
            return compatible[0]
        if len(compatible) > 1 and target.team:
            team_hits = prefer_generation(
                target, [h for h in compatible if h.team == target.team]
            )
            if len(team_hits) == 1:
                return team_hits[0]
    return None


def best_name(*players: PlayerAdp) -> tuple[str, str, str]:
    def score(p: PlayerAdp) -> tuple:
        _, first, last = name_parts(p.name)
        return (
            len(last),
            0 if is_attachable_short(p.name) else 1,
            len(first),
            1 if p.team else 0,
            1 if p.pos else 0,
            len(p.name),
        )

    best = max(players, key=score)
    team = next((p.team for p in players if p.team), "")
    pos = max((p.pos for p in players), key=lambda x: POS_PREF.get(x, 2))
    if (pos == "DST" or any(p.pos == "DST" for p in players)) and team and team in ABBR_TO_FULL:
        return ABBR_TO_FULL[team], "DST", team
    return best.name, pos or best.pos, team or best.team


def compare_lists(a: list[PlayerAdp], b: list[PlayerAdp], label_a: str, label_b: str, cap: float = 300) -> dict[str, Any]:
    idx_b = index_players(b)
    diffs: list[tuple[str, float, float, float]] = []
    for p in a:
        if p.adp > cap:
            continue
        q = match_player(p, idx_b)
        if not q or q.adp > cap:
            continue
        diffs.append((p.name, p.adp, q.adp, abs(p.adp - q.adp)))
    if not diffs:
        return {"label_a": label_a, "label_b": label_b, "overlap": 0, "mean_abs": None, "big": []}
    mean_abs = sum(d[3] for d in diffs) / len(diffs)
    big = [d for d in diffs if d[3] >= 15]
    big.sort(key=lambda x: -x[3])
    return {
        "label_a": label_a,
        "label_b": label_b,
        "overlap": len(diffs),
        "mean_abs": round(mean_abs, 2),
        "pct_big": round(100 * len(big) / len(diffs), 1),
        "big": big[:12],
        "similar": mean_abs <= 8 and (len(big) / len(diffs) <= 0.15),
    }


def pick_native(candidates: dict[str, list[PlayerAdp]], source: str) -> tuple[list[PlayerAdp], str, list[str]]:
    """candidates keys are origin tags. Return (list, origin_used, notes)."""
    notes: list[str] = []
    if len(candidates) == 1:
        origin, lst = next(iter(candidates.items()))
        return lst, origin, notes
    order = NATIVE_PRIORITY.get(source, list(candidates.keys()))
    ranked = [(o, candidates[o]) for o in order if o in candidates]
    ranked += [(o, lst) for o, lst in candidates.items() if o not in {r[0] for r in ranked}]
    if not ranked:
        return [], "", notes
    native_origin, native_list = ranked[0]
    for origin, lst in ranked[1:]:
        cmp = compare_lists(native_list, lst, native_origin, origin)
        if cmp["overlap"] == 0:
            notes.append(f"{source}: {origin} did not overlap {native_origin}; keeping {native_origin}.")
            continue
        if cmp["similar"]:
            notes.append(
                f"{source}: {origin} vs {native_origin} mean |diff| {cmp['mean_abs']} "
                f"on {cmp['overlap']} players — using native {native_origin}."
            )
        else:
            notes.append(
                f"{source}: MAJOR DIFF {origin} vs {native_origin} mean |diff| {cmp['mean_abs']} "
                f"({cmp['pct_big']}% off by 15+) on {cmp['overlap']} players. Using native {native_origin}."
            )
            for name, a, b, d in cmp["big"][:8]:
                notes.append(f"    {name}: {native_origin} {a} vs {origin} {b} (Δ{d:.0f})")
    return native_list, native_origin, notes


def consensus_avg_and_n(ranks: dict[str, float], pos: str) -> tuple[float, int]:
    """Average may ignore early K/DST ranks; # Sites is always the filled site count."""
    avg_src = ranks
    if pos in {"K", "DST"}:
        kept = {s: a for s, a in ranks.items() if a >= K_DST_MIN_CONSENSUS_ADP}
        avg_src = kept if kept else {s: max(a, K_DST_MIN_CONSENSUS_ADP) for s, a in ranks.items()}
    avg = sum(avg_src.values()) / len(avg_src)
    return round(avg, 2), len(ranks)


def merge_bucket(sources: dict[str, list[PlayerAdp]]) -> list[dict[str, Any]]:
    """Union players; compute equal-weight average ADP across sites that ranked them."""
    universe: list[PlayerAdp] = []
    for lst in sources.values():
        universe.extend(lst)
    slot_firsts = slot_full_firsts(universe)

    groups: list[list[PlayerAdp]] = []
    full_index: dict[tuple[str, str, str], int] = {}

    def group_keys(p: PlayerAdp) -> list[tuple[str, str, str]]:
        # Full-name keys only — abbreviated names attach later.
        return identity_keys(p, fuzzy=False)

    for p in universe:
        if is_attachable_short(p.name):
            continue
        gi = None
        for k in group_keys(p):
            if k in full_index:
                gi = full_index[k]
                break
        if gi is None:
            gi = len(groups)
            groups.append([])
        groups[gi].append(p)
        for k in group_keys(p):
            full_index.setdefault(k, gi)

    # Attach "J. Gibbs" / "Q. Judkins" / last-name-only rows to the unique full-name group.
    short_idx = index_players([p for g in groups for p in g])
    unmatched_shorts: list[PlayerAdp] = []
    for p in universe:
        if not is_attachable_short(p.name):
            continue
        hit = match_player(p, short_idx, slot_firsts=slot_firsts)
        if hit is None:
            unmatched_shorts.append(p)
            continue
        attached = False
        for g in groups:
            if any(x.name == hit.name and x.pos == hit.pos and x.team == hit.team for x in g):
                g.append(p)
                attached = True
                break
        if not attached:
            unmatched_shorts.append(p)
    for p in unmatched_shorts:
        groups.append([p])

    source_indexes = {src: index_players(lst) for src, lst in sources.items()}
    rows: list[dict[str, Any]] = []
    for group in groups:
        if not group:
            continue
        name, pos, team = best_name(*group)
        by_src: dict[str, float] = {}
        for src, idx in source_indexes.items():
            hit = match_player(PlayerAdp(name, pos, team, 1), idx, slot_firsts=slot_firsts)
            if hit:
                by_src[src] = hit.adp
        if not by_src:
            continue
        avg, n = consensus_avg_and_n(by_src, pos)
        rows.append(
            {
                "name": name,
                "pos": pos,
                "team": team,
                "avg": avg,
                "n": n,
                "ranks": by_src,
            }
        )
    return collapse_rows(rows, slot_firsts)


def unify_dst_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """One D/ST row per NFL team. Nicknames, 'Bills DST', and 'Buffalo Bills' collapse together."""
    others: list[dict[str, Any]] = []
    by_team: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        dst = resolve_defense(row["name"], row["pos"], row["team"])
        if not dst:
            others.append(row)
            continue
        _name, _pos, team = dst
        row["name"], row["pos"], row["team"] = dst
        by_team[team].append(row)

    merged: list[dict[str, Any]] = []
    for team, cluster in by_team.items():
        ranks: dict[str, float] = {}
        for r in cluster:
            for src, adp in r["ranks"].items():
                ranks.setdefault(src, adp)
        if not ranks:
            continue
        avg, n = consensus_avg_and_n(ranks, "DST")
        merged.append(
            {
                "name": ABBR_TO_FULL[team],
                "pos": "DST",
                "team": team,
                "avg": avg,
                "n": n,
                "ranks": ranks,
            }
        )

    out = others + merged
    out.sort(key=lambda r: (r["avg"], r["name"]))
    for i, r in enumerate(out, 1):
        r["rank"] = i
    return out


def collapse_rows(
    rows: list[dict[str, Any]],
    slot_firsts: dict[tuple[str, str, str], set[str]] | None = None,
) -> list[dict[str, Any]]:
    """Merge leftover abbreviated duplicates (A. St. Brown / Amon-Ra St. Brown)."""
    clusters: list[list[dict[str, Any]]] = []
    for row in rows:
        placed = False
        for cluster in clusters:
            head = cluster[0]
            if not slots_compatible(row["pos"], head["pos"], row["team"], head["team"]):
                continue
            teams_agree = bool(row["team"] and head["team"] and row["team"] == head["team"])
            if not last_names_compatible(row["name"], head["name"], teams_agree=teams_agree):
                continue
            if not first_names_compatible(row["name"], head["name"], teams_agree=teams_agree):
                continue
            if slot_firsts and initial_collision(
                row["name"], head["name"], row["pos"] or head["pos"], row["team"] or head["team"], slot_firsts
            ):
                continue
            cluster.append(row)
            placed = True
            break
        if not placed:
            clusters.append([row])

    collapsed: list[dict[str, Any]] = []
    for cluster in clusters:
        names = [PlayerAdp(r["name"], r["pos"], r["team"], r["avg"]) for r in cluster]
        name, pos, team = best_name(*names)
        ranks: dict[str, float] = {}
        for r in cluster:
            for src, adp in r["ranks"].items():
                ranks.setdefault(src, adp)
        if not ranks:
            continue
        avg, n = consensus_avg_and_n(ranks, pos)
        collapsed.append(
            {
                "name": name,
                "pos": pos,
                "team": team,
                "avg": avg,
                "n": n,
                "ranks": ranks,
            }
        )
    collapsed = attach_short_name_rows(collapsed)
    collapsed.sort(key=lambda r: (r["avg"], r["name"]))
    for i, r in enumerate(collapsed, 1):
        r["rank"] = i
    return unify_dst_rows(collapsed)


def attach_short_name_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Fold leftover B. Robinson rows into the lowest-ADP full name (Bijan, not Brian)."""
    longs: list[dict[str, Any]] = []
    shorts: list[dict[str, Any]] = []
    for row in rows:
        if is_attachable_short(row["name"]):
            shorts.append(row)
        else:
            longs.append(row)
    if not shorts:
        return rows
    for short in shorts:
        _, initial, last = name_parts(short["name"])
        cands: list[dict[str, Any]] = []
        for row in longs:
            if not slots_compatible(short["pos"], row["pos"], short["team"], row["team"]):
                continue
            teams_agree = bool(short["team"] and row["team"] and short["team"] == row["team"])
            if not last_names_compatible(short["name"], row["name"], teams_agree=teams_agree):
                continue
            rf = name_parts(row["name"])[1]
            if is_last_name_only(short["name"]):
                cands.append(row)
            elif len(initial) == 1 and rf.startswith(initial):
                cands.append(row)
        if not cands:
            longs.append(short)
            continue
        if short["pos"]:
            pos_cands = [c for c in cands if c["pos"] == short["pos"]]
            if pos_cands:
                cands = pos_cands
        if short["team"]:
            team_cands = [c for c in cands if c["team"] == short["team"]]
            if team_cands:
                cands = team_cands
        winner = min(cands, key=lambda r: (r["avg"], r["name"]))
        for src, adp in short["ranks"].items():
            winner["ranks"].setdefault(src, adp)
        winner["avg"], winner["n"] = consensus_avg_and_n(winner["ranks"], winner["pos"])
    return longs


def short_list_note(src: str, players: list[PlayerAdp]) -> str | None:
    n = len(players)
    if n == 0:
        return None
    if n < 200:
        return f"{src}: {n} ranked players (shorter than the typical 200–300)."
    return None


def collect_all() -> tuple[dict[str, dict[str, list[PlayerAdp]]], list[str], dict[str, list[str]]]:
    notes: list[str] = []
    shorts: dict[str, list[str]] = defaultdict(list)
    buckets: dict[str, dict[str, dict[str, list[PlayerAdp]]]] = defaultdict(lambda: defaultdict(dict))
    # buckets[bucket][source][origin] = list

    def add(bucket: str, source: str, origin: str, players: list[PlayerAdp]) -> None:
        if not players:
            return
        if "_season_" in bucket:
            players = [p for p in players if "pearsall" not in name_parts(p.name)[0]]
            if not players:
                return
        buckets[bucket][source][origin] = players
        msg = short_list_note(f"{bucket} / {source} ({origin})", players)
        if msg:
            shorts[bucket].append(msg)

    ppr_xlsx = ADP_DIR / "PPR Rankings.xlsx"
    half_xlsx = ADP_DIR / "Half PPR Rankings.xlsx"
    std_xlsx = ADP_DIR / "Standard Rankings.xlsx"
    qb2_ppr = ADP_DIR / "2QB PPR Rankings .xlsx"
    qb2_half = ADP_DIR / "2QB Half PPR Rankings.xlsx"
    qb2_std = ADP_DIR / "2QB Standard Rankings.xlsx"
    dyn_xlsx = ADP_DIR / "Dynasty Rankings.xlsx"

    # --- Redraft PPR 1QB ---
    for src, lst in parse_fp_adp(ppr_xlsx, "FP ADP (YahooSleepRTSport) ppr").items():
        add("ppr_season_1qb", src, "fp_adp", lst)
    for src, lst in parse_pff(ppr_xlsx, "PFF ppr").items():
        add("ppr_season_1qb", src, "pff_sheet", lst)
    for src, lst in parse_draftsharks(ppr_xlsx, "DraftSharkCBSESPNConsensus").items():
        add("ppr_season_1qb", src, "draftsharks", lst)
    for src, lst in parse_cbs(ppr_xlsx, "CBS").items():
        add("ppr_season_1qb", src, "cbs_sheet", lst)
    for src, lst in parse_yahoo(ppr_xlsx, "Yahoo").items():
        add("ppr_season_1qb", src, "yahoo_sheet", lst)

    # --- Redraft half PPR 1QB ---
    for src, lst in parse_fp_adp(half_xlsx, "FP ADP (YahooSleepRTSport) 12").items():
        add("half_ppr_season_1qb", src, "fp_adp", lst)
    for src, lst in parse_pff(half_xlsx, "PFF 12ppr").items():
        add("half_ppr_season_1qb", src, "pff_sheet", lst)
    for src, lst in parse_draftsharks(half_xlsx, "DraftSharkSleeperConsensusYahoo").items():
        add("half_ppr_season_1qb", src, "draftsharks", lst)
    for src, lst in parse_yahoo(half_xlsx, "Yahoo").items():
        add("half_ppr_season_1qb", src, "yahoo_sheet", lst)

    # --- Redraft Standard 1QB ---
    for src, lst in parse_fp_adp(std_xlsx, "FP ADP (YahooSleepRTSport)stand").items():
        add("standard_season_1qb", src, "fp_adp", lst)
    for src, lst in parse_pff(std_xlsx, "PFF standard").items():
        add("standard_season_1qb", src, "pff_sheet", lst)
    for src, lst in parse_draftsharks(std_xlsx, "SleeperCBSConsensus").items():
        add("standard_season_1qb", src, "draftsharks", lst)
    for src, lst in parse_cbs(std_xlsx, "CBS").items():
        add("standard_season_1qb", src, "cbs_sheet", lst)

    # --- Redraft 2QB PPR ---
    for src, lst in parse_fp_rank_list(qb2_ppr, "FP 2QB ppr").items():
        add("ppr_season_superflex", src, "fp_ecr", lst)
    for src, lst in parse_pff(qb2_ppr, "PFF 2QB ppr").items():
        add("ppr_season_superflex", src, "pff_sheet", lst)
    for src, lst in parse_draftsharks(qb2_ppr, "SleeperConsensus").items():
        add("ppr_season_superflex", src, "draftsharks", lst)

    # --- Redraft 2QB half PPR ---
    for src, lst in parse_fp_rank_list(qb2_half, "FP 2QB 12ppr").items():
        add("half_ppr_season_superflex", src, "fp_ecr", lst)
    # Sleeper publishes one Superflex ADP (`adp_2qb`), not scoring-specific lists.
    sleeper_2qb = parse_sleeper_tsv(ADP_DIR / "sleeper-2qb-adp.tsv")
    add("ppr_season_superflex", "sleeper", "sleeper_api", sleeper_2qb)
    add("half_ppr_season_superflex", "sleeper", "sleeper_api", sleeper_2qb)
    add("standard_season_superflex", "sleeper", "sleeper_api", sleeper_2qb)

    # --- Redraft 2QB Standard ---
    for src, lst in parse_fp_rank_list(qb2_std, "2QB Standard").items():
        add("standard_season_superflex", src, "fp_ecr", lst)

    # Underdog
    add("ppr_season_1qb", "underdog", "underdog_csv", parse_underdog(ADP_DIR / "Underdog 2026 PPR.csv"))
    add("half_ppr_season_1qb", "underdog", "underdog_csv", parse_underdog(ADP_DIR / "Underdog 2026 season.csv"))
    add("ppr_season_superflex", "underdog", "underdog_csv", parse_underdog(ADP_DIR / "Underdog 2QB.csv"))

    # ESPN TSV from the earlier download
    espn_tsv: dict[str, list[PlayerAdp]] = {}
    if ESPN_TSV.exists():
        espn_tsv = parse_espn_tsv(ESPN_TSV)
        add("ppr_season_1qb", "espn", "espn_tsv", espn_tsv.get("ppr", []))
        add("standard_season_1qb", "espn", "espn_tsv", espn_tsv.get("standard", []))
        add("ppr_season_superflex", "espn", "espn_tsv", espn_tsv.get("superflex", []))

    # ESPN cheat-sheet PDFs (PPR / standard / superflex / dynasty). TSV is the same
    # three redraft formats scraped from ESPN; PDF wins when both exist.
    espn_ppr = parse_espn_pdf(ADP_DIR / "ESPN PPR NFL26_CS_PPR300.pdf")
    espn_std = parse_espn_pdf(ADP_DIR / "ESPN26_CS_Non300.pdf")
    espn_sf = parse_espn_pdf(ADP_DIR / "ESPN 2QB NFL26_CS_Super.pdf")
    espn_dyn = parse_espn_pdf(ADP_DIR / "ESPN Dynasty NFL26_CS_Dyn.pdf")
    add("ppr_season_1qb", "espn", "espn_pdf", espn_ppr["overall"])
    add("standard_season_1qb", "espn", "espn_pdf", espn_std["overall"])
    add("ppr_season_superflex", "espn", "espn_pdf", espn_sf["overall"])
    add("ppr_dynasty_1qb", "espn", "espn_pdf", espn_dyn["overall"])

    # Dynasty
    for src, lst in parse_fp_rank_list(dyn_xlsx, "FP PPR Dynasty").items():
        add("ppr_dynasty_1qb", src, "fp_adp", lst)
    for src, lst in parse_pff(dyn_xlsx, "PFF PPR DYN").items():
        add("ppr_dynasty_1qb", src, "pff_sheet", lst)
    for src, lst in parse_sleeper_ds(dyn_xlsx, "Sleeper PPR (DYN)").items():
        add("ppr_dynasty_1qb", src, "sleeper_sheet", lst)

    for src, lst in parse_sleeper_ds(dyn_xlsx, "Sleeper 12ppr (DYN)").items():
        add("half_ppr_dynasty_1qb", src, "sleeper_sheet", lst)

    for src, lst in parse_sleeper_ds(dyn_xlsx, "Sleeper Standard (DYN)").items():
        add("standard_dynasty_1qb", src, "sleeper_sheet", lst)

    espn_tsv_deep = parse_espn_tsv(ESPN_TSV, cap=None) if ESPN_TSV.exists() else {}
    ppr_redraft = merge_espn_overall(espn_tsv_deep.get("ppr", []), espn_ppr["overall"])
    std_redraft = merge_espn_overall(espn_tsv_deep.get("standard", []), espn_std["overall"])
    sf_redraft = merge_espn_overall(espn_tsv_deep.get("superflex", []), espn_sf["overall"])
    add(
        "standard_season_superflex",
        "espn",
        "espn_sf_blend",
        derive_sf_espn_from_1qb(espn_std["overall"], espn_sf["overall"]),
    )

    def names_in(bucket: str) -> list[str]:
        out: list[str] = []
        for origins in buckets[bucket].values():
            for lst in origins.values():
                out.extend(p.name for p in lst)
        return out

    # ESPN dynasty PDF stops ~239. Fill remaining 1QB board names from the
    # scoring-specific redraft ESPN list (PPR TSV stands in for half PPR).
    add(
        "ppr_dynasty_1qb",
        "espn",
        "espn_dynasty_fill",
        fill_espn_from_redraft(espn_dyn["overall"], ppr_redraft, names_in("ppr_dynasty_1qb")),
    )
    add(
        "half_ppr_dynasty_1qb",
        "espn",
        "espn_dynasty_fill",
        fill_espn_from_redraft(espn_dyn["overall"], ppr_redraft, names_in("half_ppr_dynasty_1qb")),
    )
    add(
        "standard_dynasty_1qb",
        "espn",
        "espn_dynasty_fill",
        fill_espn_from_redraft(espn_dyn["overall"], std_redraft, names_in("standard_dynasty_1qb")),
    )

    for src, lst in parse_sleeper_ds(dyn_xlsx, "Sleeper 2QB PPR").items():
        add("ppr_dynasty_superflex", src, "sleeper_sheet", lst)
    for src, lst in parse_pff(dyn_xlsx, "2QB (DYN)").items():
        add("ppr_dynasty_superflex", src, "pff_sheet", lst)

    for src, lst in parse_fp_rank_list(dyn_xlsx, "FP rookies (DYN)").items():
        add("ppr_dynasty_1qb_rookies", src, "fp_adp", lst)
    for src, lst in parse_pff(dyn_xlsx, "PFF rookies (DYN)", rookie_rank=True).items():
        add("ppr_dynasty_1qb_rookies", src, "pff_sheet", lst)
    for src, lst in parse_sleeper_ds(dyn_xlsx, "Sleeper rookies (DYN)", prefer_rank=True).items():
        add("ppr_dynasty_1qb_rookies", src, "sleeper_sheet", lst)
    add(
        "ppr_dynasty_1qb_rookies",
        "yahoo",
        "yahoo_list",
        parse_sleeper_tsv(ADP_DIR / "yahoo-dynasty-1qb-rookies.tsv"),
    )

    for src, lst in parse_sleeper_ds(dyn_xlsx, "Sleeper 2QB rookies (DYN)", prefer_rank=True).items():
        add("ppr_dynasty_superflex_rookies", src, "sleeper_sheet", lst)
    for src, lst in parse_pff(dyn_xlsx, "PFF 2QB rookies (DYN)", rookie_rank=True).items():
        add("ppr_dynasty_superflex_rookies", src, "pff_sheet", lst)
    add(
        "ppr_dynasty_superflex_rookies",
        "yahoo",
        "yahoo_list",
        parse_sleeper_tsv(ADP_DIR / "yahoo-dynasty-superflex-rookies.tsv"),
    )

    # ESPN dynasty PDF only lists ~60 rookies. Backfill ESPN rookie ranks from
    # redraft PPR (1QB) / superflex (2QB) order: first rookie = 1, second = 2, …
    known_rookies = load_rookie_names_from_files()
    known_rookies.extend(p.name for p in espn_dyn["rookies"])

    def matches_name(name: str, pool: list[str]) -> bool:
        return any(names_are_same_player(name, other) for other in pool)

    # Dynasty/redraft overall boards are veterans plus the top rookies. Use them
    # only to reject Sleeper/FP "rookie" rows that are actually returning players.
    veterans = [
        p.name
        for lst in (espn_dyn["overall"], espn_ppr["overall"], espn_std["overall"], espn_sf["overall"])
        for p in lst
    ]
    extras: list[str] = []
    for bucket in ("ppr_dynasty_1qb_rookies", "ppr_dynasty_superflex_rookies"):
        for origins in buckets[bucket].values():
            for lst in origins.values():
                for p in lst:
                    if matches_name(p.name, known_rookies):
                        continue
                    if matches_name(p.name, veterans):
                        continue
                    extras.append(p.name)
    rookie_names = known_rookies + extras
    espn_rook_1qb = derive_rookie_ranks(ppr_redraft, rookie_names)
    espn_rook_sf = append_missing_rookies(
        derive_rookie_ranks(sf_redraft, rookie_names),
        espn_rook_1qb,
    )
    add("ppr_dynasty_1qb_rookies", "espn", "espn_ppr_order", espn_rook_1qb)
    add("ppr_dynasty_1qb_rookies", "espn", "espn_pdf", espn_dyn["rookies"])
    add("ppr_dynasty_superflex_rookies", "espn", "espn_ppr_order", espn_rook_sf)

    yahoo_sf = parse_numbered_name_list(ADP_DIR / "yahoo-superflex-redraft.txt")
    add("ppr_season_superflex", "yahoo", "yahoo_list", yahoo_sf)
    add("half_ppr_season_superflex", "yahoo", "yahoo_list", yahoo_sf)
    add("standard_season_superflex", "yahoo", "yahoo_list", yahoo_sf)

    yahoo_dyn_1qb = parse_sleeper_tsv(ADP_DIR / "yahoo-dynasty-1qb.tsv")
    add("ppr_dynasty_1qb", "yahoo", "yahoo_list", yahoo_dyn_1qb)
    add("half_ppr_dynasty_1qb", "yahoo", "yahoo_list", yahoo_dyn_1qb)
    add("standard_dynasty_1qb", "yahoo", "yahoo_list", yahoo_dyn_1qb)
    yahoo_sf_tagged = overlay_identity(
        yahoo_sf,
        yahoo_dyn_1qb,
        parse_sleeper_tsv(ADP_DIR / "yahoo-dynasty-1qb-rookies.tsv"),
        parse_sleeper_tsv(ADP_DIR / "yahoo-dynasty-superflex-rookies.tsv"),
    )
    yahoo_dyn_sf = derive_sf_espn_from_1qb(yahoo_dyn_1qb, yahoo_sf_tagged)
    add("ppr_dynasty_superflex", "yahoo", "yahoo_sf_blend", yahoo_dyn_sf)
    add("half_ppr_dynasty_superflex", "yahoo", "yahoo_sf_blend", yahoo_dyn_sf)
    add("standard_dynasty_superflex", "yahoo", "yahoo_sf_blend", yahoo_dyn_sf)

    espn_dyn_1qb = parse_espn_labeled_list(ADP_DIR / "espn-dynasty-1qb-top240.txt")
    espn_dyn_sf = parse_espn_labeled_list(ADP_DIR / "espn-dynasty-superflex.txt")
    espn_dyn_sf_blend = derive_sf_espn_from_1qb(espn_dyn_1qb, espn_dyn_sf)
    add("ppr_dynasty_superflex", "espn", "espn_sf_blend", espn_dyn_sf_blend)
    add("half_ppr_dynasty_superflex", "espn", "espn_sf_blend", espn_dyn_sf_blend)
    add("standard_dynasty_superflex", "espn", "espn_sf_blend", espn_dyn_sf_blend)

    def copy_source(src_bucket: str, dst_bucket: str, source: str) -> None:
        origins = buckets[src_bucket].get(source) or {}
        for origin, lst in origins.items():
            add(dst_bucket, source, origin, lst)

    # ESPN publishes one ranking for PPR and half PPR (redraft + dynasty).
    copy_source("ppr_season_1qb", "half_ppr_season_1qb", "espn")
    copy_source("ppr_season_superflex", "half_ppr_season_superflex", "espn")
    copy_source("ppr_dynasty_1qb", "half_ppr_dynasty_1qb", "espn")
    copy_source("ppr_dynasty_superflex", "half_ppr_dynasty_superflex", "espn")

    # Yahoo has no Standard redraft 1QB board; half PPR is the closest published list.
    copy_source("half_ppr_season_1qb", "standard_season_1qb", "yahoo")

    # Sleeper dynasty 2QB half/standard uses the PPR 2QB board.
    copy_source("ppr_dynasty_superflex", "half_ppr_dynasty_superflex", "sleeper")
    copy_source("ppr_dynasty_superflex", "standard_dynasty_superflex", "sleeper")
    copy_source("ppr_dynasty_superflex", "half_ppr_dynasty_superflex", "pff")
    copy_source("ppr_dynasty_superflex", "standard_dynasty_superflex", "pff")

    # Resolve native copies
    resolved: dict[str, dict[str, list[PlayerAdp]]] = {}
    for bucket, src_map in buckets.items():
        resolved[bucket] = {}
        for source, origins in src_map.items():
            lst, origin, src_notes = pick_native(origins, source)
            resolved[bucket][source] = lst
            notes.extend([f"[{bucket}] {m}" for m in src_notes])
            notes.append(f"[{bucket}] {source}: {len(lst)} players from {origin}")

    return resolved, notes, shorts


BUCKET_TITLES = {
    "ppr_season_1qb": "PPR · Redraft · 1QB",
    "half_ppr_season_1qb": "Half PPR · Redraft · 1QB",
    "standard_season_1qb": "Standard · Redraft · 1QB",
    "ppr_season_superflex": "PPR · Redraft · 2QB",
    "half_ppr_season_superflex": "Half PPR · Redraft · 2QB",
    "standard_season_superflex": "Standard · Redraft · 2QB",
    "ppr_dynasty_1qb": "PPR · Dynasty · 1QB",
    "half_ppr_dynasty_1qb": "Half PPR · Dynasty · 1QB",
    "standard_dynasty_1qb": "Standard · Dynasty · 1QB",
    "ppr_dynasty_superflex": "PPR · Dynasty · 2QB",
    "half_ppr_dynasty_superflex": "Half PPR · Dynasty · 2QB",
    "standard_dynasty_superflex": "Standard · Dynasty · 2QB",
    "ppr_dynasty_1qb_rookies": "PPR · Dynasty · 1QB · Rookies",
    "ppr_dynasty_superflex_rookies": "PPR · Dynasty · 2QB · Rookies",
}

# Manual ADP patches when a player moved after the source files were exported.
SITE_OVERRIDES: dict[str, dict[str, dict[str, Any]]] = {
    "ppr_season_1qb": {
        "stefon diggs": {
            "team": "WAS",
            "ranks": {
                "cbs": 123.0,
                "sleeper": 116.0,
                "fantasypros": 113.0,
                "rtsports": 109.0,
                "fantrax": 81.0,
            },
        }
    },
    "half_ppr_season_1qb": {
        "stefon diggs": {
            "team": "WAS",
            "ranks": {
                "yahoo": 117.0,
                "sleeper": 118.0,
                "fantasypros": 115.0,
                "rtsports": 114.0,
            },
        }
    },
    "standard_season_1qb": {
        "stefon diggs": {
            "team": "WAS",
            "ranks": {
                "sleeper": 130.0,
                "fantasypros": 122.0,
                "rtsports": 114.0,
            },
        }
    },
}


def apply_site_overrides(bucket: str, rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    spec = SITE_OVERRIDES.get(bucket)
    if not spec:
        return rows
    changed = False
    for row in rows:
        patch = spec.get(name_parts(row["name"])[0])
        if not patch:
            continue
        changed = True
        if patch.get("team"):
            row["team"] = patch["team"]
        for src, adp in (patch.get("ranks") or {}).items():
            if adp is None:
                row["ranks"].pop(src, None)
            else:
                row["ranks"][src] = float(adp)
        if row["ranks"]:
            row["avg"], row["n"] = consensus_avg_and_n(row["ranks"], row["pos"])
    if not changed:
        return rows
    rows.sort(key=lambda r: (r["avg"], r["name"]))
    for i, r in enumerate(rows, 1):
        r["rank"] = i
    return rows


SOURCE_ORDER = [
    "espn",
    "cbs",
    "yahoo",
    "sleeper",
    "fantasypros",
    "pff",
    "underdog",
    "rtsports",
    "fantrax",
    "draftsharks",
    "nfl",
]


def blend_espn_from_peer_ranks(rows: list[dict[str, Any]], *names: str) -> None:
    """Replace ESPN's rank with the mean of the other sites on this board."""
    want = {name_parts(n)[0] for n in names}
    for row in rows:
        if name_parts(row["name"])[0] not in want:
            continue
        peers = [v for src, v in row["ranks"].items() if src != "espn" and v is not None]
        if not peers:
            continue
        row["ranks"]["espn"] = round(sum(peers) / len(peers), 1)
        row["avg"], row["n"] = consensus_avg_and_n(row["ranks"], row["pos"])


def fill_sleeper_from_fp(rows: list[dict[str, Any]]) -> None:
    """Sleeper dynasty sheets often omit names FantasyPros already ranked. Copy the FP cell."""
    for row in rows:
        if row["ranks"].get("sleeper") is None and row["ranks"].get("fantasypros") is not None:
            row["ranks"]["sleeper"] = row["ranks"]["fantasypros"]
            row["avg"], row["n"] = consensus_avg_and_n(row["ranks"], row["pos"])


def load_identity_catalog() -> dict[str, tuple[str, str]]:
    catalog: dict[str, tuple[str, str]] = dict(KNOWN_SLOTS)
    files = [
        ROOT / "rankings" / "generated" / "rookies_dynasty_1qb.csv",
        ROOT / "rankings" / "generated" / "rookies_dynasty_superflex.csv",
    ]
    for path in files:
        if not path.exists():
            continue
        with path.open(encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                name = (row.get("name") or row.get("Player") or "").strip()
                pos = canon_pos(row.get("position") or row.get("Pos") or "")
                team = canon_team(row.get("team") or row.get("Team") or "")
                n = name_parts(name)[0]
                if n and (pos or team):
                    prev = catalog.get(n, ("", ""))
                    catalog[n] = (pos or prev[0], team or prev[1])
    return catalog


def fill_missing_identity(all_rows: dict[str, list[dict[str, Any]]]) -> None:
    catalog = load_identity_catalog()
    for rows in all_rows.values():
        for row in rows:
            n = name_parts(row["name"])[0]
            if not n:
                continue
            if row["pos"] and row["team"]:
                prev = catalog.get(n, ("", ""))
                catalog[n] = (row["pos"] or prev[0], row["team"] or prev[1])
    for rows in all_rows.values():
        for row in rows:
            n = name_parts(row["name"])[0]
            slot = catalog.get(n)
            if not slot:
                for kn, cand in catalog.items():
                    if names_are_same_player(row["name"], kn):
                        slot = cand
                        break
            if not slot:
                continue
            pos, team = slot
            if not row["pos"] and pos:
                row["pos"] = pos
            if not row["team"] and team:
                row["team"] = team


def fill_missing_dst(
    all_rows: dict[str, list[dict[str, Any]]],
    src_cols_by_bucket: dict[str, list[str]],
) -> None:
    for bucket, ref_bucket in DST_FILL_FROM.items():
        rows = all_rows.get(bucket)
        ref_rows = all_rows.get(ref_bucket)
        if not rows or not ref_rows:
            continue
        src_cols = src_cols_by_bucket.get(bucket) or []
        have = {r["team"] for r in rows if r["pos"] == "DST" and r["team"]}
        ref_by_team = {r["team"]: r for r in ref_rows if r["pos"] == "DST" and r["team"]}
        for team in NFL_DST_TEAMS:
            if team in have:
                continue
            ref = ref_by_team.get(team)
            if not ref:
                continue
            ranks = {s: adp for s, adp in ref["ranks"].items() if s in src_cols}
            if not ranks:
                ranks = {src_cols[0]: ref["avg"]} if src_cols else {}
            if not ranks:
                continue
            avg, n = consensus_avg_and_n(ranks, "DST")
            rows.append(
                {
                    "name": ABBR_TO_FULL.get(team, team),
                    "pos": "DST",
                    "team": team,
                    "avg": avg,
                    "n": n,
                    "ranks": ranks,
                }
            )
            have.add(team)
        rows.sort(key=lambda r: (r["avg"], r["name"]))
        for i, r in enumerate(rows, 1):
            r["rank"] = i


def write_outputs(
    resolved: dict[str, dict[str, list[PlayerAdp]]],
    notes: list[str],
    shorts: dict[str, list[str]],
) -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    avg_fill = PatternFill("solid", fgColor="FFF2CC")

    catalog: dict[str, Any] = {"buckets": {}}
    merged: dict[str, list[dict[str, Any]]] = {}
    src_cols_by_bucket: dict[str, list[str]] = {}

    for bucket, sources in resolved.items():
        rows = apply_site_overrides(bucket, merge_bucket(sources))
        if bucket == "ppr_dynasty_1qb":
            fill_sleeper_from_fp(rows)
        if bucket == "ppr_dynasty_1qb_rookies":
            # ESPN 1QB rookie order comes from redraft PPR and buries QBs.
            # Mendoza is the exception: use the other sites' 1QB rookie mean.
            blend_espn_from_peer_ranks(rows, "Fernando Mendoza")
        src_cols = [s for s in SOURCE_ORDER if s in sources]
        src_cols += [s for s in sources if s not in src_cols]
        merged[bucket] = rows
        src_cols_by_bucket[bucket] = src_cols

    fill_missing_identity(merged)
    fill_missing_dst(merged, src_cols_by_bucket)

    for bucket, sources in resolved.items():
        rows = merged[bucket]
        src_cols = src_cols_by_bucket[bucket]
        rows.sort(key=lambda r: (r["avg"], r["name"]))
        for i, r in enumerate(rows, 1):
            r["rank"] = i

        headers = ["Rank", "Player", "Pos", "Team", "Avg ADP", "# Sites"] + [SOURCE_LABELS.get(s, s) for s in src_cols]

        csv_path = OUT_DIR / f"{bucket}.csv"
        with csv_path.open("w", encoding="utf-8", newline="") as f:
            w = csv.writer(f)
            w.writerow(headers)
            for r in rows:
                line = [r["rank"], r["name"], r["pos"], r["team"], r["avg"], r["n"]]
                for s in src_cols:
                    v = r["ranks"].get(s)
                    line.append("" if v is None else v)
                w.writerow(line)

        ws = wb.create_sheet(BUCKET_TITLES.get(bucket, bucket)[:31])
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(1, col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        ws.cell(1, 5).fill = PatternFill("solid", fgColor="C65911")
        for r in rows:
            line = [r["rank"], r["name"], r["pos"], r["team"], r["avg"], r["n"]]
            for s in src_cols:
                v = r["ranks"].get(s)
                line.append(None if v is None else v)
            ws.append(line)
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
        ws.freeze_panes = "A2"
        widths = [8, 28, 8, 8, 12, 10] + [12] * len(src_cols)
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
            for cell in row:
                cell.fill = avg_fill
                cell.number_format = "0.00"

        catalog["buckets"][bucket] = {
            "title": BUCKET_TITLES.get(bucket, bucket),
            "sources": src_cols,
            "players": [
                {
                    "rank": r["rank"],
                    "name": r["name"],
                    "pos": r["pos"],
                    "team": r["team"],
                    "avg": r["avg"],
                    "n": r["n"],
                    "ranks": r["ranks"],
                }
                for r in rows
            ],
        }

    wb.save(OUT_DIR / "master.xlsx")
    (OUT_DIR / "catalog.json").write_text(json.dumps(catalog), encoding="utf-8")

    report: list[str] = []
    report.append("ADP consensus parse report")
    report.append("")
    for bucket, sources in resolved.items():
        title = BUCKET_TITLES.get(bucket, bucket)
        rows = catalog["buckets"][bucket]["players"]
        report.append(f"== {title} ({bucket}) ==")
        report.append(f"  players: {len(rows)}  sources: {', '.join(SOURCE_LABELS.get(s, s) for s in sources)}")
        dst_rows = [r for r in rows if r["pos"] == "DST"]
        dst_teams = {r["team"] for r in dst_rows if r["team"]}
        report.append(f"  DST: {len(dst_rows)} rows / {len(dst_teams)} teams (max 32)")
        if len(dst_rows) > 32:
            report.append("  WARNING: more than 32 DST rows after unify")
        extra = [r["name"] for r in dst_rows if not r["team"]]
        if extra:
            report.append(f"  DST missing team: {', '.join(extra[:12])}")
        for src, lst in sources.items():
            report.append(f"    {SOURCE_LABELS.get(src, src)}: {len(lst)}")
        if shorts.get(bucket):
            report.append("  short lists:")
            for msg in shorts[bucket]:
                report.append(f"    - {msg}")
        report.append("")
    report.append("Native vs duplicate copies")
    report.append("")
    report.extend(notes)
    (OUT_DIR / "_report.txt").write_text("\n".join(report), encoding="utf-8")
    print("\n".join(report[:80]))
    print(f"\nWrote {OUT_DIR}")


def main() -> None:
    resolved, notes, shorts = collect_all()
    write_outputs(resolved, notes, shorts)


if __name__ == "__main__":
    main()
