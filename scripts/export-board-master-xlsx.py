#!/usr/bin/env python3
"""Turn rankings/adp-sources/board-master.json into board-master.xlsx."""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "rankings" / "adp-sources" / "board-master.json"
OUT = ROOT / "rankings" / "adp-sources" / "board-master.xlsx"

HEADERS = ["Player", "Pos", "Team", "Avg ADP", "Consensus", "ESPN", "Yahoo", "Sleeper"]
GAP_HEADERS = [
    "Format",
    "Player",
    "Pos",
    "Team",
    "Consensus",
    "ESPN",
    "Yahoo",
    "Sleeper",
    "Gap",
    "Highest (worst)",
    "Lowest (best)",
]
GAP_MIN = 8
GAP_CAP = 300
GAP_LIMIT = 50


def site_gaps(rows: list[dict], *, cap: int = GAP_CAP, min_gap: int = GAP_MIN, limit: int = GAP_LIMIT) -> list[dict]:
    """Early ranks count more: a 12-spot spread at pick 20 outranks a 40-spot spread at 250."""
    out: list[dict] = []
    for row in rows:
        consensus = int(row["consensus"])
        if consensus > cap:
            continue
        ranks = {
            "ESPN": int(row["espn"]),
            "Yahoo": int(row["yahoo"]),
            "Sleeper": int(row["sleeper"]),
        }
        gap = max(ranks.values()) - min(ranks.values())
        if gap < min_gap:
            continue
        high = max(ranks, key=ranks.get)
        low = min(ranks, key=ranks.get)
        out.append(
            {
                **row,
                "gap": gap,
                "high": f"{high} {ranks[high]}",
                "low": f"{low} {ranks[low]}",
                "weight": gap / (max(consensus, 1) ** 0.55),
            }
        )
    out.sort(key=lambda r: (-r["weight"], -r["gap"], r["consensus"]))
    return out[:limit]


def write_header(ws, headers, header_font, header_fill) -> None:
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(1, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")


def main() -> None:
    payload = json.loads(SRC.read_text(encoding="utf-8"))
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    avg_fill = PatternFill("solid", fgColor="FFF2CC")
    gap_fill = PatternFill("solid", fgColor="F4B183")
    widths = [28, 8, 8, 12, 12, 12, 12, 12]

    cover = wb.create_sheet("How to read", 0)
    cover["A1"] = "Board master"
    cover["A1"].font = Font(bold=True, size=14)
    cover["A3"] = (
        "Each format sheet is Consensus / ESPN / Yahoo / Sleeper in published-board "
        "order (site names first, consensus tail after). Avg ADP is the multi-site "
        "average; tail-only players use Consensus so the cell is never empty. "
        "Team is the NFL abbreviation: LA is written LAR (Rams), Chargers stay LAC, "
        "D/ST use the club code, and unsigned/blank pool teams show as FA. "
        "Gaps sheets list top-300 consensus players with the largest early-weighted "
        "spreads across the four boards. A 12-spot gap in the first 50 ranks higher "
        "than a 40-spot gap near pick 250."
    )
    cover["A3"].alignment = Alignment(wrap_text=True, vertical="top")
    cover.column_dimensions["A"].width = 110
    cover.row_dimensions[3].height = 110

    for sheet in payload["sheets"]:
        title = sheet["title"][:31]
        ws = wb.create_sheet(title)
        ws.append(HEADERS)
        for col in range(1, len(HEADERS) + 1):
            cell = ws.cell(1, col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        ws.cell(1, 4).fill = PatternFill("solid", fgColor="C65911")
        for row in sheet["rows"]:
            ws.append(
                [
                    row["name"],
                    row["pos"],
                    row["team"],
                    row["avg"],
                    row["consensus"],
                    row["espn"],
                    row["yahoo"],
                    row["sleeper"],
                ]
            )
        last = ws.max_row
        ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{last}"
        ws.freeze_panes = "A2"
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        for cell_row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
            for cell in cell_row:
                cell.fill = avg_fill
                cell.number_format = "0.00"
        for col in range(5, 9):
            for cell_row in ws.iter_rows(min_row=2, min_col=col, max_col=col):
                for cell in cell_row:
                    cell.number_format = "0"
        print(f"{title}: {last - 1} rows")

    combined: list[tuple[str, dict]] = []
    ppr_1qb: list[dict] = []
    for sheet in payload["sheets"]:
        gaps = site_gaps(sheet["rows"])
        for row in gaps:
            combined.append((sheet["title"], row))
        if sheet.get("key") == "ppr_season_1qb":
            ppr_1qb = gaps

    def write_gap_sheet(title: str, items: list[tuple[str, dict]], *, include_format: bool) -> None:
        ws = wb.create_sheet(title[:31], 1)
        headers = GAP_HEADERS if include_format else [h for h in GAP_HEADERS if h != "Format"]
        write_header(ws, headers, header_font, header_fill)
        ws.cell(1, headers.index("Gap") + 1).fill = gap_fill
        for fmt, row in items:
            line = [
                row["name"],
                row["pos"],
                row["team"],
                row["consensus"],
                row["espn"],
                row["yahoo"],
                row["sleeper"],
                row["gap"],
                row["high"],
                row["low"],
            ]
            if include_format:
                line.insert(0, fmt)
            ws.append(line)
        last = max(ws.max_row, 1)
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last}"
        ws.freeze_panes = "A2"
        gap_widths = ([36] if include_format else []) + [28, 8, 8, 12, 12, 12, 12, 8, 18, 18]
        for i, width in enumerate(gap_widths[: len(headers)], 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        print(f"{title}: {max(last - 1, 0)} rows")

    if combined:
        write_gap_sheet("Gaps · all formats", combined, include_format=True)
    if ppr_1qb:
        write_gap_sheet("Gaps · PPR 1QB", [("", row) for row in ppr_1qb], include_format=False)

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
